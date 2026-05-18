# python_processor.py
# CAD Column Inspector Pro - Python Processor
# Copyright (c) 2026 Tran Xuan An

import sys
import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class PolygonProcessor:
    """Process polygon data from AutoLISP and export to Excel"""
    
    def __init__(self, json_file):
        self.json_file = json_file
        self.polygons = []
        self.output_folder = os.path.dirname(json_file)
        
    def load_data(self):
        """Load polygon data from JSON file"""
        try:
            with open(self.json_file, 'r') as f:
                data = json.load(f)
            
            # Handle both single object (old format) and array (new format)
            if isinstance(data, list):
                self.polygons = data
            else:
                # Old format: single polygon object
                self.polygons = [data]
            
            print(f"✓ Loaded {len(self.polygons)} polygon(s) from: {self.json_file}")
            return True
        except Exception as e:
            print(f"✗ Error loading data: {e}")
            return False
    
    def process(self):
        """Process polygon data"""
        if not self.polygons:
            print("✗ No data to process")
            return False
        
        print("\n=== Processing Polygon Data ===")
        for idx, poly in enumerate(self.polygons, start=1):
            print(f"\nPolygon #{idx}:")
            print(f"  Name: {poly.get('name', 'Unknown')}")
            print(f"  Type: {poly.get('type')}")
            print(f"  Point Count: {poly.get('point_count')}")
            print(f"  Area: {poly.get('area'):.2f} sq units")
            print(f"  Centroid: {poly.get('centroid')}")
            print(f"  Drawing: {poly.get('drawing')}")
        
        return True
    
    def export_to_excel(self):
        """Export data to Excel file"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_folder,
                f"Column_Inspector_{timestamp}.xlsx"
            )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Polygon Data"
            
            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            row = 1
            
            # Process each polygon
            for poly_idx, polygon in enumerate(self.polygons, start=1):
                # Section header
                if poly_idx > 1:
                    row += 2  # Add space between polygons
                
                ws.cell(row=row, column=1).value = f"POLYGON #{poly_idx}"
                ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FF0000")
                row += 1
                
                # Property headers
                headers = ["Property", "Value"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                row += 1
                
                # Write data
                ws.cell(row=row, column=1).value = "Name"
                ws.cell(row=row, column=2).value = polygon.get('name', 'Unknown')
                ws.cell(row=row, column=2).font = Font(bold=True, color="0000FF")
                row += 1
                
                ws.cell(row=row, column=1).value = "Type"
                ws.cell(row=row, column=2).value = polygon.get('type', 'N/A')
                row += 1
                
                ws.cell(row=row, column=1).value = "Point Count"
                ws.cell(row=row, column=2).value = polygon.get('point_count', 0)
                row += 1
                
                ws.cell(row=row, column=1).value = "Area (sq units)"
                ws.cell(row=row, column=2).value = round(polygon.get('area', 0), 2)
                row += 1
                
                centroid = polygon.get('centroid', [0, 0])
                ws.cell(row=row, column=1).value = "Centroid X"
                ws.cell(row=row, column=2).value = round(centroid[0], 2)
                row += 1
                
                ws.cell(row=row, column=1).value = "Centroid Y"
                ws.cell(row=row, column=2).value = round(centroid[1], 2)
                row += 1
                
                ws.cell(row=row, column=1).value = "Drawing"
                ws.cell(row=row, column=2).value = polygon.get('drawing', 'N/A')
                row += 1
                
                ws.cell(row=row, column=1).value = "Timestamp"
                ws.cell(row=row, column=2).value = polygon.get('timestamp', 'N/A')
                row += 2
                
                # Points section
                ws.cell(row=row, column=1).value = "Points:"
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                # Points header
                point_headers = ["Point #", "X", "Y", "Z"]
                for col, header in enumerate(point_headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                row += 1
                
                # Points data
                points = polygon.get('points', [])
                for idx, point in enumerate(points, start=1):
                    ws.cell(row=row, column=1).value = idx
                    ws.cell(row=row, column=2).value = round(point[0], 2)
                    ws.cell(row=row, column=3).value = round(point[1], 2)
                    ws.cell(row=row, column=4).value = round(point[2], 2) if len(point) > 2 else 0
                    row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            # Save workbook
            wb.save(output_file)
            print(f"\n✓ Excel file created with {len(self.polygons)} polygon(s): {output_file}")
            
            return output_file
            
        except Exception as e:
            print(f"✗ Error creating Excel file: {e}")
            return None


def main():
    """Main entry point"""
    print("=" * 60)
    print("   CAD Column Inspector Pro - Python Processor")
    print("=" * 60)
    
    # Check arguments
    if len(sys.argv) < 2:
        print("\nUsage: python python_processor.py <json_file>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    
    # Check file exists
    if not os.path.exists(json_file):
        print(f"✗ Error: File not found: {json_file}")
        sys.exit(1)
    
    # Process
    processor = PolygonProcessor(json_file)
    
    if not processor.load_data():
        sys.exit(1)
    
    if not processor.process():
        sys.exit(1)
    
    output_file = processor.export_to_excel()
    
    if output_file:
        print("\n✓ Processing completed successfully!")
        print("=" * 60)
        sys.exit(0)
    else:
        print("\n✗ Processing failed!")
        sys.exit(1)


if __name__ == "__main__":
    main()
